import sys
import json
import os
import openpyxl
from openpyxl.styles import Font

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QTreeView, QHeaderView, QStyleFactory,
    QPushButton, QWidget, QVBoxLayout, QHBoxLayout, QInputDialog,
    QMessageBox, QSplitter, QLabel, QAction, QFileDialog, QLineEdit
)
from PyQt5.QtGui import QStandardItemModel, QStandardItem, QIcon
from PyQt5.QtCore import Qt

# Custom model to make parents read-only
class ReadOnlyParentModel(QStandardItemModel):
    def flags(self, index):
        default_flags = super().flags(index)
        if not index.isValid(): return Qt.NoItemFlags
        item = self.itemFromIndex(index)
        if item and item.hasChildren():
            return default_flags & ~Qt.ItemIsEditable
        else:
            return default_flags

# Panel 1: Project Management (Unchanged)
class ProjectPanel(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        layout = QVBoxLayout(self); layout.setContentsMargins(0,0,0,0)
        title = QLabel("<b>Project Management (Parents are Read-Only)</b>"); title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)
        self.tree_view = QTreeView(); layout.addWidget(self.tree_view)
        self.model = ReadOnlyParentModel()
        self.model.setHorizontalHeaderLabels(['Name', 'Status', 'Assignee'])
        self.populate_data()
        self.tree_view.setModel(self.model)
        self.configure_tree_view()
        self.setup_buttons(layout)

    def setup_buttons(self, layout):
        button_layout = QHBoxLayout()
        add_col_button = QPushButton("Add Column")
        add_col_button.clicked.connect(self.add_column)
        button_layout.addStretch()
        button_layout.addWidget(add_col_button)
        layout.addLayout(button_layout)

    def add_column(self):
        column_name, ok = QInputDialog.getText(self, "Add Column", "Enter new column header:")
        if ok and column_name:
            current_count = self.model.columnCount()
            self.model.setColumnCount(current_count + 1)
            self.model.setHeaderData(current_count, Qt.Horizontal, column_name)
            def populate_new_column(parent_item):
                for row in range(parent_item.rowCount()):
                    new_item = QStandardItem(""); new_item.setEditable(True)
                    parent_item.setChild(row, current_count, new_item)
                    child_parent = parent_item.child(row, 0)
                    if child_parent: populate_new_column(child_parent)
            populate_new_column(self.model.invisibleRootItem())
    
    def configure_tree_view(self):
        self.tree_view.expandAll()
        self.tree_view.header().setSectionResizeMode(0, QHeaderView.Stretch)
        self.tree_view.setAlternatingRowColors(True)

    def populate_data(self):
        root = self.model.invisibleRootItem()
        p1_name = QStandardItem("Project Alpha"); p1_name.setEditable(True)
        root.appendRow([p1_name, QStandardItem("In Progress"), QStandardItem("Team A")])
        t1_name = QStandardItem("Task 1.1: Design UI"); t1_name.setEditable(True)
        p1_name.appendRow([t1_name, QStandardItem("Completed"), QStandardItem("Alice")])


# Panel 2: Sequential Steps (Updated Delete Logic)
class StepsPanel(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.step_counter = 1
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        title = QLabel("<b>Sequential Steps (Parents are Read-Only)</b>")
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)
        self.tree_view = QTreeView()
        layout.addWidget(self.tree_view)
        self.model = ReadOnlyParentModel()
        self.model.setHorizontalHeaderLabels(['Task Name'])
        self.populate_data()
        self.tree_view.setModel(self.model)
        self.configure_tree_view()
        self.setup_buttons(layout)

    def setup_buttons(self, layout):
        button_layout = QHBoxLayout()
        add_step_button = QPushButton("Add Step")
        add_step_button.setIcon(QIcon.fromTheme("list-add"))
        add_step_button.clicked.connect(self.add_step_column)
        delete_step_button = QPushButton("Delete Last Step")
        delete_step_button.setIcon(QIcon.fromTheme("list-remove"))
        delete_step_button.clicked.connect(self.delete_step_column)
        button_layout.addStretch()
        button_layout.addWidget(add_step_button)
        button_layout.addWidget(delete_step_button)
        layout.addLayout(button_layout)

    def delete_step_column(self):
        current_column_count = self.model.columnCount()
        if current_column_count <= 1:
            QMessageBox.warning(self, "Delete Step", "There are no step columns to delete.")
            return
        self.model.removeColumn(current_column_count - 1)
        if self.step_counter > 1:
            self.step_counter -= 1

    def add_step_column(self):
        column_name = f"Step {self.step_counter}"
        new_column_index = self.model.columnCount()
        self.model.setColumnCount(new_column_index + 1)
        self.model.setHeaderData(new_column_index, Qt.Horizontal, column_name)
        def populate_new_column(parent_item):
            for row in range(parent_item.rowCount()):
                new_item = QStandardItem(""); new_item.setEditable(True)
                parent_item.setChild(row, new_column_index, new_item)
                child_parent = parent_item.child(row, 0)
                if child_parent: populate_new_column(child_parent)
        populate_new_column(self.model.invisibleRootItem())
        self.step_counter += 1

    def configure_tree_view(self):
        self.tree_view.expandAll()
        self.tree_view.header().setSectionResizeMode(0, QHeaderView.Stretch)
        self.tree_view.setAlternatingRowColors(True)

    def populate_data(self):
        root = self.model.invisibleRootItem()
        phase1 = QStandardItem("Phase 1: Planning"); phase1.setEditable(True)
        root.appendRow(phase1)
        sub1 = QStandardItem("Sub-task A: Research"); sub1.setEditable(True)
        sub2 = QStandardItem("Sub-task A: DD"); sub1.setEditable(True)
        phase1.appendRow(sub1)
        phase1.appendRow(sub2)
        phase2 = QStandardItem("Phase 2: Execution"); phase2.setEditable(True)
        root.appendRow(phase2)



# Main Window (MODIFIED)
# (Keep all your existing imports at the top of the file)
# Make sure to add this one if it's not already there for openpyxl > 2.4
from openpyxl.styles import Font, Alignment 

# ... (Keep ReadOnlyParentModel, ProjectPanel, and StepsPanel classes as they are) ...

# Main Window (CORRECTED)
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("PyQt5 Dual Tree Table App")
        self.setGeometry(300, 200, 1200, 800)
        main_container = QWidget(); main_layout = QVBoxLayout(main_container)
        self.setup_menu()
        splitter = QSplitter(Qt.Vertical)
        self.project_panel = ProjectPanel(); self.steps_panel = StepsPanel()
        splitter.addWidget(self.project_panel); splitter.addWidget(self.steps_panel)
        splitter.setSizes([400, 400])
        main_layout.addWidget(splitter)
        self.setup_global_buttons(main_layout)
        self.setCentralWidget(main_container)

    def setup_global_buttons(self, layout):
        # (This function is unchanged from your original code)
        button_layout = QHBoxLayout()
        button_layout.addWidget(QLabel("Filename:"))
        self.filename_input = QLineEdit("my_project.json")
        button_layout.addWidget(self.filename_input)
        save_button = QPushButton("Save")
        save_button.setIcon(QIcon.fromTheme("document-save"))
        save_button.clicked.connect(self.quick_save_state)
        button_layout.addWidget(save_button)
        export_xlsx_button = QPushButton("Export to XLSX")
        export_xlsx_button.setIcon(QIcon.fromTheme("document-export"))
        export_xlsx_button.clicked.connect(self.export_to_xlsx)
        button_layout.addWidget(export_xlsx_button)
        button_layout.addStretch()
        layout.addLayout(button_layout)
    
    def _get_max_depth(self, parent_item):
        """
        FIXED: Recursively finds the number of levels in the tree.
        A tree with only top-level items has a depth of 1.
        A tree with a child has a depth of 2, etc.
        """
        if not parent_item.hasChildren():
            return 0
        
        max_child_depth = 0
        for r in range(parent_item.rowCount()):
            child_item = parent_item.child(r, 0)
            # The depth of this branch is 1 (for the child itself) plus the depth of its own subtree.
            depth = 1 + self._get_max_depth(child_item)
            if depth > max_child_depth:
                max_child_depth = depth
        return max_child_depth

    def _write_model_to_sheet_merged(self, model, sheet):
        """
        Writes a QStandardItemModel to an openpyxl sheet with a merged, hierarchical structure.
        """
        bold_font = Font(bold=True)
        center_align = Alignment(vertical='center', horizontal='left', wrap_text=True)

        def _count_branch_rows(item):
            if not item or not item.hasChildren():
                return 1
            count = 0
            for i in range(item.rowCount()):
                count += _count_branch_rows(item.child(i, 0))
            return count

        def _recursive_write_and_merge(parent_item, start_row, depth):
            current_row = start_row
            col_offset = depth * model.columnCount()

            for r in range(parent_item.rowCount()):
                item_for_this_row = parent_item.child(r, 0)
                row_span = _count_branch_rows(item_for_this_row)
                
                for c in range(model.columnCount()):
                    cell_item = parent_item.child(r, c)
                    text = cell_item.text() if cell_item else ""
                    
                    cell = sheet.cell(row=current_row, column=col_offset + c + 1, value=text)
                    cell.alignment = center_align
                    
                    if item_for_this_row.hasChildren():
                        cell.font = bold_font
                    
                    if row_span > 1:
                        sheet.merge_cells(
                            start_row=current_row,
                            start_column=col_offset + c + 1,
                            end_row=current_row + row_span - 1,
                            end_column=col_offset + c + 1
                        )

                if item_for_this_row.hasChildren():
                    _recursive_write_and_merge(item_for_this_row, current_row, depth + 1)
                
                current_row += row_span

        # --- Main logic for _write_model_to_sheet_merged ---
        # 1. Determine headers based on max depth
        base_headers = [model.headerData(i, Qt.Horizontal) for i in range(model.columnCount())]
        # FIXED: Correctly calculate the number of levels and generate headers.
        num_levels = self._get_max_depth(model.invisibleRootItem())
        if num_levels == 0 and model.rowCount() > 0: # Handle case with only top-level items
            num_levels = 1
        
        all_headers = []
        for i in range(num_levels):
            all_headers.extend(base_headers)
        
        if all_headers:
            sheet.append(all_headers)
            for cell in sheet[1]:
                cell.font = bold_font

        # 2. Start the recursive writing process
        _recursive_write_and_merge(model.invisibleRootItem(), start_row=2, depth=0)
        
        # 3. Auto-adjust column widths
        for i, column_cells in enumerate(sheet.columns):
            # Use the header text length as a minimum width
            header_text = all_headers[i] if i < len(all_headers) else ''
            max_length = len(header_text)
            column = column_cells[0].column_letter

            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            # Add a little padding, but cap the max width
            adjusted_width = min((max_length + 2) * 1.2, 50)
            sheet.column_dimensions[column].width = adjusted_width

    def export_to_xlsx(self):
        options = QFileDialog.Options()
        filePath, _ = QFileDialog.getSaveFileName(self, "Export to XLSX", "", "Excel Files (*.xlsx);;All Files (*)", options=options)
        if not filePath: return
        try:
            wb = openpyxl.Workbook()
            
            # Export Project Panel
            sheet1 = wb.active
            sheet1.title = "Project Management"
            self._write_model_to_sheet_merged(self.project_panel.model, sheet1)

            # Export Steps Panel
            sheet2 = wb.create_sheet(title="Sequential Steps")
            self._write_model_to_sheet_merged(self.steps_panel.model, sheet2)
            
            wb.save(filePath)
            self.statusBar().showMessage(f"Successfully exported to {filePath}", 5000)
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Could not export file:\n{e}")

    # (The rest of the MainWindow class: setup_menu, _get_app_state_data, quick_save_state, save_state_as, load_state, serialize_model, deserialize_model remains unchanged)
    # ... PASTE THE REST OF YOUR ORIGINAL MainWindow CODE HERE ...
    def setup_menu(self):
        menubar = self.menuBar()
        file_menu = menubar.addMenu('&File')
        save_action = QAction('&Save', self); save_action.setShortcut('Ctrl+S'); save_action.triggered.connect(self.quick_save_state)
        file_menu.addAction(save_action)
        save_as_action = QAction('Save &As...', self); save_as_action.setShortcut('Ctrl+Shift+S'); save_as_action.triggered.connect(self.save_state_as)
        file_menu.addAction(save_as_action)
        load_action = QAction('&Load State...', self); load_action.setShortcut('Ctrl+O'); load_action.triggered.connect(self.load_state)
        file_menu.addAction(load_action)
        file_menu.addSeparator()
        exit_action = QAction('&Exit', self); exit_action.setShortcut('Ctrl+Q'); exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)
        self.statusBar()

    def _get_app_state_data(self):
        project_model = self.project_panel.model
        project_headers = [project_model.headerData(i, Qt.Horizontal) for i in range(project_model.columnCount())]
        project_tree = self.serialize_model(project_model)
        steps_model = self.steps_panel.model
        steps_headers = [steps_model.headerData(i, Qt.Horizontal) for i in range(steps_model.columnCount())]
        steps_tree = self.serialize_model(steps_model)
        return {
            'project_panel': {'headers': project_headers, 'tree': project_tree},
            'steps_panel': {'headers': steps_headers, 'tree': steps_tree, 'step_counter': self.steps_panel.step_counter}
        }

    def quick_save_state(self):
        filename = self.filename_input.text().strip()
        if not filename: QMessageBox.warning(self, "Save Error", "Filename cannot be empty."); return
        if not filename.lower().endswith('.json'): filename += '.json'
        app_state = self._get_app_state_data()
        try:
            with open(filename, 'w') as f: json.dump(app_state, f, indent=4)
            self.statusBar().showMessage(f"State saved to {filename}", 5000)
        except Exception as e:
            QMessageBox.critical(self, "Error Saving", f"Could not save file:\n{e}")

    def save_state_as(self):
        options = QFileDialog.Options(); filePath, _ = QFileDialog.getSaveFileName(self, "Save State As...", "", "JSON Files (*.json);;All Files (*)", options=options)
        if not filePath: return
        app_state = self._get_app_state_data()
        try:
            with open(filePath, 'w') as f: json.dump(app_state, f, indent=4)
            self.statusBar().showMessage(f"State saved to {filePath}", 5000)
            self.filename_input.setText(os.path.basename(filePath))
        except Exception as e:
            QMessageBox.critical(self, "Error Saving", f"Could not save file:\n{e}")

    def load_state(self):
        options = QFileDialog.Options(); filePath, _ = QFileDialog.getOpenFileName(self, "Load State", "", "JSON Files (*.json);;All Files (*)", options=options)
        if not filePath: return
        try:
            with open(filePath, 'r') as f: app_state = json.load(f)
        except Exception as e:
            QMessageBox.critical(self, "Error Loading", f"Could not load or parse file:\n{e}"); return
        project_data = app_state.get('project_panel', {}); project_model = self.project_panel.model; project_model.clear()
        project_model.setHorizontalHeaderLabels(project_data.get('headers', []))
        self.deserialize_model(project_model, project_data.get('tree', []), project_model.invisibleRootItem())
        self.project_panel.tree_view.expandAll()
        steps_data = app_state.get('steps_panel', {}); steps_model = self.steps_panel.model; steps_model.clear()
        steps_model.setHorizontalHeaderLabels(steps_data.get('headers', []))
        self.deserialize_model(steps_model, steps_data.get('tree', []), steps_model.invisibleRootItem())
        self.steps_panel.tree_view.expandAll()
        self.steps_panel.step_counter = steps_data.get('step_counter', 1)
        self.statusBar().showMessage(f"State loaded from {filePath}", 5000)
        self.filename_input.setText(os.path.basename(filePath))

    def serialize_model(self, model):
        def recurse(parent_item):
            nodes = []
            for r in range(parent_item.rowCount()):
                node_data = []
                for c in range(model.columnCount()):
                    item = parent_item.child(r, c)
                    text = item.text() if item else ""
                    node_data.append(text)
                parent_for_children = parent_item.child(r, 0)
                node = {'data': node_data, 'children': recurse(parent_for_children)}
                nodes.append(node)
            return nodes
        return recurse(model.invisibleRootItem())

    def deserialize_model(self, model, tree_data, parent_item):
        for node in tree_data:
            row_items = []
            for text in node['data']:
                item = QStandardItem(text); item.setEditable(True)
                row_items.append(item)
            parent_item.appendRow(row_items)
            new_parent = parent_item.child(parent_item.rowCount() - 1, 0)
            if node['children']: self.deserialize_model(model, node['children'], new_parent)

# ... (The if __name__ == '__main__': block remains unchanged)
if __name__ == '__main__':
    app = QApplication(sys.argv)
    app.setStyle(QStyleFactory.create("Fusion"))
    main_window = MainWindow()
    main_window.show()
    sys.exit(app.exec_())
